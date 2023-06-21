import gym
import torch as th
import torch.nn as nn
import highway_env
from stable_baselines3.td3_ver1.td3_ver1 import TD3_ver1
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import numpy as np
import time
from openpyxl import Workbook
from highway_env.envs.myhighway_utils import MPC, Model
from highway_env import utils
import openpyxl
if __name__ == '__main__':

    env = gym.make("myhighway-v5")
    env.configure({
        "observation": {
            "type": "MPCObservation_v0",
            "features": ['x', 'y', 'heading', 'speed', 'target_y'],

            # "type": "KinematicDecision4",

        },
        "action": {
                "type": "ContinuousAction",
                'steering_range': (-np.pi / 10, np.pi / 10),
                'acceleration_range': (-5, 5),
                # "dynamical": True, # 먼저 학습 해 보고 잘되면 나중에 더하
            },
        "reward_speed_slope": [0, 1],
        "collision_reward": -10,  # The reward received when colliding with a vehicle.
        "target_lane_reward": 0.5,
        "high_speed_reward": 0.5,
        "lanes_count": 4,
        "duration" : 1000, # 속력 20m/s이면 1200m가지 진행 (60sec)
        "stop_speed" : 10,
        "vehicles_count": 50,
        "random_road": False, #몇 차선인지 결정
        'road_static': True,  #미리 정의해놓은 global path & initial position?
        'save_positions': False, # global path & initial position 저장?
    })
    j = 0
    iteration = 1
    wait_time = 0

    info_dict = {}
    # position, action ... oscillation &  uncontrollable .... log 기록 (trajectory)
    write_wb = Workbook()
    model_path = 'save_path/path.xlsx'
    model = Model()
    # model.p['target_v', 30]
    # model.p['dt', 0.1]
    mpc = MPC(model)
    dt = 0.1
    action = np.zeros(2)
    write_wb = openpyxl.load_workbook(model_path)
    for i in range(iteration):
        done = False
        obs = env.reset()

        position_x = []
        position_y = []
        angular_velocity = []
        acceleration = []
        datas = []

        # x0 = obs['map'].reshape(len(env.config['observation']['features']),1)
        x0 = np.concatenate([obs['map'].reshape(-1, 1), obs['surroundings'].reshape(-1, 1)], axis=0)

        mpc.x0 = x0
        mpc.set_initial_guess()
        u0 = 0
        flag = 0
        while not done:
            # print("{:.2f}, {}, 감속 : {}, 지점 : {:.2f}, ttc : {}".format(env.controlled_vehicles[0].speed, \
            #                                                           obs['decision'][0], env.decelerate_flag, \
            #                                                           env.controlled_vehicles[0].position[0], \
            #                                                           obs['ttc']))
            tic = time.time()
            output = u0
            u0 = mpc.make_step(x0)

            toc = time.time()
            u0 = u0.reshape(1,2)[0]
            u0[0] = utils.lmap(u0[0], env.config['action']['acceleration_range'],[-1, 1])
            u0[1] = utils.lmap(u0[1], env.config['action']['steering_range'],[-1, 1])
            obs, reward, done, info = env.step(u0)
            # obs, reward, done, info = env.step([0,0])
            # obs, reward, done, info = env.step([1,1])
            # print(mpc.data.success)
            # x0 = obs['map'].reshape(len(env.config['observation']['features']),1)
            x0 = np.concatenate([obs['map'].reshape(-1, 1), obs['surroundings'].reshape(-1, 1)], axis=0)

            if not mpc.data.success[-1][0]:
                print('fail Step : {}'.format(env.steps))
                # print(mpc.store_solver_stats)
                print(mpc.solver_stats['success'], mpc.solver_stats['return_status'])
                flag = 1
                # done = True
            if flag and mpc.data.success[-1][0]:

                # print(mpc.store_solver_stats)
                print(mpc.solver_stats['success'], mpc.solver_stats['return_status'])
                flag = 0

            datas.append([env.controlled_vehicles[0].position[0], env.controlled_vehicles[0].position[1], \
                          env.controlled_vehicles[0].action['steering'], \
                          env.controlled_vehicles[0].action['acceleration'], \
                          env.controlled_vehicles[0].speed, \
                          mpc.data.success[-1][0],\
                          obs['map'][4], env.controlled_vehicles[0].heading,obs['surroundings'][3], \
                          u0[0],u0[1]])

            if done:
                j = 1
                np.save('save_path/MPCpath_' + str(env.config["lanes_count"]) + '.npy', datas)

            # env.render('human_fast')

            if j == 0:
                time.sleep(wait_time)
                j = 1

            # time.sleep(0.3)

        write_ws = write_wb.create_sheet("Trial{}".format(i))
        # col_names = ['pos x', 'pos y', 'steering', 'acc', 'speed', 'solver', 'target y', 'global x', 'global y']
        col_names = ['pos x', 'pos y', 'steering', 'acc', 'speed', 'solver', 'target y', 'heading','distance','input1','input2']
        write_ws.append(col_names)
        for c in range(len(datas)):
            write_ws.append(datas[c])

        # path_x = env.global_path[0]
        # path_x = np.dstack((path_x - 1, path_x))
        # path_x = path_x.reshape(-1)
        # path_y = env.global_path[1] * 4
        # path_y = np.dstack((path_y, np.append(path_y[:-1], 0)))
        # path_y = path_y.reshape(-1)
        #
        # for c in range(len(path_x)):
        #     write_ws.cell(row=2 + c, column=len(col_names) - 1, value=path_x[c])
        #     write_ws.cell(row=2 + c, column=len(col_names), value=path_y[c])

    write_wb.save(model_path + '.xlsx')
    env.close()


