import gym
import torch as th
import torch.nn as nn
import highway_env
from stable_baselines3.td3_ver1.td3_ver1 import TD3_ver1
from stable_baselines3 import TD3
from stable_baselines3 import HerReplayBuffer
from stable_baselines3.td3_ver1.td3_ver2 import TD3_ver2

from stable_baselines3.td3_ver1.td3_ver3 import TD3_ver3

from stable_baselines3.td3_ver1.td3_ver4 import TD3_ver4

from stable_baselines3.common.vec_env import VecVideoRecorder, DummyVecEnv
from stable_baselines3.common.torch_layers import BaseFeaturesExtractor
from stable_baselines3.common.buffers import DecisionReplayBuffer
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import numpy as np

# kinematics observation with decision replay buffer

import gym
import torch as th
import torch.nn as nn
import highway_env
from stable_baselines3.td3_ver1.td3_ver1 import TD3_ver1
from stable_baselines3 import TD3
from stable_baselines3 import HerReplayBuffer
from stable_baselines3.td3_ver1.td3_ver2 import TD3_ver2

from stable_baselines3.td3_ver1.td3_ver3 import TD3_ver3

from stable_baselines3.td3_ver1.td3_ver4 import TD3_ver4

import openpyxl
import pandas as pd
import time
from stable_baselines3.common.vec_env import VecVideoRecorder, DummyVecEnv
from stable_baselines3.common.torch_layers import BaseFeaturesExtractor
from stable_baselines3.common.buffers import DecisionReplayBuffer
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import numpy as np

# kinematics observation with decision replay buffer


if __name__ == '__main__':

    env = gym.make("refhighway-v0")
    env.configure({
        "observation": {
            "type": "FrontDist1",
        },
        "action": {
            "type": "ContinuousAction",
            'steering_range': (-np.pi / 10, np.pi / 10),
            'acceleration_range': (-5, 5),
            "dynamical": False,  # 먼저 학습 해 보고 잘되면 나중에 더하
        },
        "collision_reward": -5,  # The reward received when colliding with a vehicle.
        "duration": 1000,  # 속력 20m/s이면 1200m가지 진행 (60sec)
        "stop_speed": 15,
        "lanes_count": 4,
        # Fixted target 하고싶으면,
        # 'mpc_path': '/Users/cml/highway-env/scripts/MPC',
        # # 학습 끝나고 테스트 할 때 아래 주석 풀기!
        # "random_road": False,  # 몇 차선인지 결정
        # 'road_static': True,  # 미리 정의해놓은 global path & initial position?
        # 'save_positions': False,  # global path & initial position 저장?
        # 'train': False,
    })
    env.reset()

    model = TD3_ver4("MultiInputPolicy",
                env,
                verbose=1,
                buffer_size=1_000_000,
                train_freq=(1, "step"),
                learning_rate=1e-4, # actor : 1e-4, critic : 1e-5
                batch_size=100,
                save_path='/Users/cml/highway-env/scripts/MPC/new_reward2/model{}_{}',
                save_interval=5e3,
                tensorboard_log="MPC/")

    Train = True
    if Train:
        # model_path = '/Users/cml/highway-env/scripts/MPC/RL_MPC_fixed_2/model88_191'
        # model = TD3_ver4.load(model_path, env=env)

        model.learn(total_timesteps=int(5e6),tb_log_name='RL_MPC', log_interval=20,)
        model.save("/Users/cml/highway-env/scripts/MPC/model")
    else:
        save_path = "/Users/cml/highway-env/scripts/MPC/RL_MPC_fixed_3/RLMPC.xlsx"
        model_path = "/Users/cml/highway-env/scripts/MPC/new_reward2/model261_301"
        model = TD3_ver4.load(model_path, env=env)
        iteration = 1
        j = 0
        wait_time = 0

        info_dict = {}
        # position, action ... oscillation &  uncontrollable .... log 기록 (trajectory)
        write_wb = openpyxl.load_workbook(save_path)
        # write_wb = openpyxl.Workbook()
        style = 'action'

        for i in range(iteration):
            done = False
            obs = env.reset()
            steer = 0
            position_x = []
            position_y = []
            angular_velocity = []
            acceleration = []
            datas = []
            while not done:


                action, _ = model.predict(obs, deterministic=True)

                obs, reward, done, info = env.step(action)

                close_v = env.road.close_vehicles_to(env.controlled_vehicles[0], 200.0, count=4, see_behind=True,
                                                     sort=True)
                temp = pd.DataFrame.from_records([close_v[0].to_dict()])[['x', 'y', 'speed']]
                temp = temp.append(pd.DataFrame.from_records(
                    [v.to_dict()
                     for v in close_v[-5 + 2:]])[['x', 'y', 'speed']],
                                   ignore_index=True)
                close = temp.values.copy()
                vid = np.zeros([4, 1])
                for i in range(4):
                    vid[i] = int(str(close_v[i]).split('#')[1].split(':')[0])
                    # vid = vid.append(str(close_v[i]))
                    # print(str(close_v[i]))
                close_v_data = np.concatenate((vid, close), axis=1)
                close_v_data = close_v_data.flatten()
                new_data = [env.controlled_vehicles[0].position[0], env.controlled_vehicles[0].position[1], \
                            env.controlled_vehicles[0].action['steering'], \
                            env.controlled_vehicles[0].action['acceleration'], \
                            env.controlled_vehicles[0].speed, \
                            # obs['decision'][0],\
                            env.decision, \
                            env.target_lane * 4, \
                            reward, env.ref_path[env.steps][0],env.ref_path[env.steps][1],env.ref_path[env.steps][2],env.ref_path[env.steps][3]]
                # new_data.extend(close_v_data) # 주변 차량 데이터 더하는 부분


                datas.append(new_data)

                if done:
                    j = 1
                env.render('human_fast')

                if j == 0:
                    time.sleep(wait_time)
                    j = 1

                # time.sleep(0.3)

            ## 경로 에러 저장하는 코드
            write_ws = write_wb.create_sheet("Trial{}".format(style))
            col_names = ['pos x', 'pos y', 'steering', 'acc', 'speed', 'decision', 'target y','reward', 'mpc_x','mpc_y',
                         'mpc_steer','mpc_acc']
            # col_names = ['pos x', 'pos y', 'steering', 'acc', 'speed', 'decision', 'target y', 'reward',
            #              'id1', 'xo1', 'yo1', 'vo1', 'id2', 'xo2', 'yo2', 'vo2', 'id3', 'xo3', 'yo3', 'vo3', 'id4',
            #              'xo4', 'yo4', 'vo4']

            write_ws.append(col_names)
            for c in range(len(datas)):
                write_ws.append(datas[c])
            # global_path = env.global_path.tolist()
            path_x = env.global_path[0]
            path_x = np.dstack((path_x - 1, path_x))
            path_x = path_x.reshape(-1)
            path_y = env.global_path[1] * 4
            path_y = np.dstack((path_y, np.append(path_y[:-1], 0)))
            path_y = path_y.reshape(-1)

            # for c in range(len(path_x)):
            #     write_ws.cell(row=2 + c, column=len(col_names) - 1, value=path_x[c])
            #     write_ws.cell(row=2+c, column=len(col_names), value=path_y[c])

        write_wb.save(save_path)
        env.close()



#
# if __name__ == '__main__':
#     Train = True
#     env = gym.make("refhighway-v0")
#     env.configure({
#         "observation": {
#             "type": "FrontDist1",
#             # "type": "NoDecision1",
#         },
#         "action": {
#             "type": "ContinuousAction",
#             'steering_range': (-np.pi / 10, np.pi / 10),
#             'acceleration_range': (-5, 5),
#             "dynamical": False,  # 먼저 학습 해 보고 잘되면 나중에 더하
#         },
#         "collision_reward": -5,  # The reward received when colliding with a vehicle.
#         "duration": 1000,  # 속력 20m/s이면 1200m가지 진행 (60sec)
#         "random_road": False, #몇 차선인지 결정
#         'road_static': True,  #미리 정의해놓은 global path & initial position?
#         'save_positions': False, # global path & initial position 저장?
#         'mpc_path': '/Users/cml/highway-env/scripts/MPC',
#     })
#     env.reset()
#
#     model = TD3_ver4("MultiInputPolicy",
#                 env,
#                 verbose=1,
#                 buffer_size=1_000_000,
#                 train_freq=(1, "step"),
#                 # replay_buffer_class=HerReplayBuffer,
#                 # max_episode_length = 1000,
#                 learning_rate=1e-4, # actor : 1e-4, critic : 1e-5
#                 batch_size=100,
#                 save_path='/Users/cml/highway-env/scripts/MPC/new_reward2/model{}_{}',
#                 save_interval=5e3,
#                 tensorboard_log="MPC/")
#     if Train:
#         # model_path = '/Users/cml/highway-env/scripts/highway_td3_decisionbuffer/TD3_v2_1step_update_2/model68_792'
#         # model = TD3_ver4.load(model_path, env=env)
#
#         model.learn(total_timesteps=int(5e6),tb_log_name='RL_MPC_fixed', log_interval=20,)
#         model.save("/Users/cml/highway-env/scripts/MPC/new_reward2/model")
