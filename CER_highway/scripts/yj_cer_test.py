import time

import gym
import openpyxl
import pandas as pd
import torch as th
import torch.nn as nn
import highway_env
from stable_baselines3 import TD3
from stable_baselines3.td3_ver1.td3_ver1 import TD3_ver1
from stable_baselines3.td3_ver1.td3_ver2 import TD3_ver2

from stable_baselines3.td3_ver1.td3_ver4 import TD3_ver4
from stable_baselines3.common.vec_env import VecVideoRecorder, DummyVecEnv
from stable_baselines3.common.torch_layers import BaseFeaturesExtractor
from openpyxl import Workbook
from collections import defaultdict
import numpy as np
import warnings

warnings.simplefilter(action='ignore', category=FutureWarning)

from stable_baselines3.td3_ver1.td3_ver3 import TD3_ver3



if __name__ == '__main__':

    Train = True
    env = gym.make("myhighway-v5")

    env.configure({
        "observation": {
            # "type": "KinematicDecision3",
            'type': "FrontDist1",
            # "type": "NoDecision1",
            # 'type': "Simple"

        },
        "action": {
            "type": "ContinuousAction",
            'steering_range': (-np.pi / 30, np.pi / 30),
            "dynamical": False,  # 먼저 학습 해 보고 잘되면 나중에 더하
        },
        "duration": 1000,
        "vehicles_count": 50,
        "lanes_count": 4,
        "reward_speed_slope": [0, 1],
        "target_lane_reward": 0.2,
        "high_speed_reward": 0.7,
        "jerk_v_reward": 0.1,
        "random_road": False, #몇 차선인지 결정
        'road_static': True,  #미리 정의해놓은 global path & initial position?
        'save_positions': False, # global path & initial position 저장?
    })

    if env.config['road_static']:
        env.config['global_path'] = np.load('save_positions/save_global_' + str(env.config["lanes_count"]) + '.npy')
        env.config['road_vehicles'] = np.load('save_positions/save_positions_' + str(env.config["lanes_count"]) + '.npy',
                                         allow_pickle=True)
    env.reset()

    # model = TD3_ver1("MultiInputPolicy",
    model=TD3_ver4("MultiInputPolicy",
    # model=TD3("MultiInputPolicy",
                     env,
                     verbose=1,
                     learning_rate=1e-4,  # actor : 1e-4, critic : 1e-5
                     save_path='/Users/cml/highway-env/scripts/highway_td3_decisionbuffer/new_reward/model{}_{}',
                     save_interval=5e4,
                     tensorboard_log="highway_td3_decisionbuffer/")

    model_path = "trained_model/model_cer1.zip"
    # model_path = "trained_model/model_cer2.zip"


    save_path = "save_path/path.xlsx"
    model = TD3_ver4.load(model_path, env=env)
    j = 0
    iteration = 1

    wait_time = 0

    info_dict = {}
    # position, action ... oscillation &  uncontrollable .... log 기록 (trajectory)
    write_wb = openpyxl.load_workbook(save_path)
    # write_wb = openpyxl.Workbook()
    style = 'action'

    for i in range(iteration):
        done = False
        obs = env.reset()
        steer = 0
        position_x = []
        position_y = []
        angular_velocity = []
        acceleration = []
        datas = []
        while not done:

            # print("{:.2f}, {}, 감속 : {}, 지점 : {:.2f}, ttc : {}".format(env.controlled_vehicles[0].speed, \
            #                                                               env.decision, env.decelerate_flag, \
            #                                                               env.controlled_vehicles[0].position[0],\
            #                                                               obs['ttc']))

            action, _ = model.predict(obs, deterministic=True)


            obs, reward, done, info = env.step(action)


            close_v = env.road.close_vehicles_to(env.controlled_vehicles[0], 200.0, count=4, see_behind=True,
                                       sort=True)
            temp = pd.DataFrame.from_records([close_v[0].to_dict()])[['x', 'y', 'speed']]
            temp = temp.append(pd.DataFrame.from_records(
                [v.to_dict()
                 for v in close_v[-5 + 2:]])[['x', 'y', 'speed']],
                               ignore_index=True)
            close = temp.values.copy()
            vid = np.zeros([4,1])
            for i in range(4):
                vid[i] = int(str(close_v[i]).split('#')[1].split(':')[0])
                # vid = vid.append(str(close_v[i]))
                # print(str(close_v[i]))
            close_v_data = np.concatenate((vid, close), axis=1)
            close_v_data = close_v_data.flatten()
            new_data = [env.controlled_vehicles[0].position[0], env.controlled_vehicles[0].position[1], \
                          env.controlled_vehicles[0].action['steering'], \
                          env.controlled_vehicles[0].action['acceleration'], \
                          env.controlled_vehicles[0].speed, \
                          # obs['decision'][0],\
                          env.decision,\
                          env.target_lane*4,\
                          reward]
            new_data.extend(close_v_data)
            datas.append(new_data)


            if done:
                j=1
            env.render('human_fast')

            if j == 0:
                time.sleep(wait_time)
                j = 1

            # time.sleep(0.3)

        ## 경로 에러 저장하는 코드



        write_ws = write_wb.create_sheet("Trial{}".format(style))
        # col_names = ['pos x', 'pos y', 'steering', 'acc', 'speed', 'decision', 'target y','reward', 'global x', 'global y']
        col_names = ['pos x', 'pos y', 'steering', 'acc', 'speed', 'decision', 'target y', 'reward',
                     'id1','xo1','yo1','vo1','id2','xo2','yo2','vo2','id3','xo3','yo3','vo3','id4','xo4','yo4','vo4']

        write_ws.append(col_names)
        for c in range(len(datas)):
            write_ws.append(datas[c])
        # global_path = env.global_path.tolist()
        path_x = env.global_path[0]
        path_x = np.dstack((path_x-1, path_x))
        path_x = path_x.reshape(-1)
        path_y = env.global_path[1]*4
        path_y = np.dstack((path_y, np.append(path_y[:-1], 0)))
        path_y = path_y.reshape(-1)

        # for c in range(len(path_x)):
        #     write_ws.cell(row=2 + c, column=len(col_names) - 1, value=path_x[c])
        #     write_ws.cell(row=2+c, column=len(col_names), value=path_y[c])




    write_wb.save(save_path)
    env.close()

