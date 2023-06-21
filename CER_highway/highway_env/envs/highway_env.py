import math

import numpy as np
import random
from highway_env.envs.myhighway_utils import emap, front_dist, front_speed
from gym.envs.registration import register
from typing import List, Tuple, Optional, Callable
from highway_env import utils
from highway_env.envs.common.abstract import AbstractEnv
from highway_env.envs.common.action import Action
from highway_env.road.road import Road, RoadNetwork, Road_Static
from highway_env.utils import near_split
from highway_env.vehicle.controller import ControlledVehicle
from highway_env.vehicle.kinematics import Vehicle
from highway_env.envs.common.observation import MPCObservation_Simple,MPCObservation
from highway_env.envs.myhighway_utils import MPC, Model
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import copy
class HighwayEnv(AbstractEnv):
    """
    A highway driving environment.

    The vehicle is driving on a straight highway with several lanes, and is rewarded for reaching a high speed,
    staying on the rightmost lanes and avoiding collisions.
    """

    @classmethod
    def default_config(cls) -> dict:
        config = super().default_config()
        config.update({
            "observation": {
                "type": "Kinematics"
            },
            "action": {
                "type": "ContinuousAction",
                # "dynamical": True,
            },
            "lanes_count": 4,
            "vehicles_count": 50,
            "controlled_vehicles": 1,
            "initial_lane_id": None,
            "duration": 40,  # [s]
            "ego_spacing": 2,
            "vehicles_density": 1,
            "collision_reward": -1,    # The reward received when colliding with a vehicle.
            "right_lane_reward": 0.1,  # The reward received when driving on the right-most lanes, linearly mapped to
                                       # zero for other lanes.
            "high_speed_reward": 0.4,  # The reward received when driving at full speed, linearly mapped to zero for
                                       # lower speeds according to config["reward_speed_range"].
            "lane_change_reward": 0,   # The reward received at each lane change action.
            "reward_speed_range": [20, 30],
            "offroad_terminal": True
        })
        return config

    def _reset(self) -> None:
        self._create_road()
        self._create_vehicles()

        self.trajectory = []

    def _create_road(self) -> None:
        """Create a road composed of straight adjacent lanes."""
        self.road = Road(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                         np_random=self.np_random, record_history=self.config["show_trajectories"])

    def _create_vehicles(self) -> None:
        """Create some new random vehicles of a given type, and add them on the road."""
        other_vehicles_type = utils.class_from_path(self.config["other_vehicles_type"])
        other_per_controlled = near_split(self.config["vehicles_count"], num_bins=self.config["controlled_vehicles"])

        self.controlled_vehicles = []
        for others in other_per_controlled:
            vehicle = Vehicle.create_random(
                self.road,
                speed=25,
                lane_id=self.config["initial_lane_id"],
                spacing=self.config["ego_spacing"]
            )
            vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
            self.controlled_vehicles.append(vehicle)
            self.road.vehicles.append(vehicle)

            for _ in range(others):
                vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
                vehicle.randomize_behavior()
                self.road.vehicles.append(vehicle)

    def _reward(self, action: Action) -> float:
        """
        The reward is defined to foster driving at high speed, on the rightmost lanes, and to avoid collisions.
        :param action: the last action performed
        :return: the corresponding reward
        """
        neighbours = self.road.network.all_side_lanes(self.vehicle.lane_index)
        lane = self.vehicle.target_lane_index[2] if isinstance(self.vehicle, ControlledVehicle) \
            else self.vehicle.lane_index[2]
        scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [0, 1])
        reward = \
            + self.config["collision_reward"] * self.vehicle.crashed \
            + self.config["right_lane_reward"] * lane / max(len(neighbours) - 1, 1) \
            + self.config["high_speed_reward"] * np.clip(scaled_speed, 0, 1)
        reward = utils.lmap(reward,
                          [self.config["collision_reward"],
                           self.config["high_speed_reward"] + self.config["right_lane_reward"]],
                          [0, 1])
        reward = 0 if not self.vehicle.on_road else reward
        return reward

    def _is_terminal(self) -> bool:
        """The episode is over if the ego vehicle crashed or the time is out."""
        return self.vehicle.crashed or \
            self.steps >= self.config["duration"] or \
            (self.config["offroad_terminal"] and not self.vehicle.on_road)

    def _cost(self, action: int) -> float:
        """The cost signal is the occurrence of collision."""
        return float(self.vehicle.crashed) or \
            self.steps >= self.config["duration"] or \
            (self.config["offroad_terminal"] and not self.vehicle.on_road)


class HighwayEnvFast(HighwayEnv):
    """
    A variant of highway-v0 with faster execution:
        - lower simulation frequency
        - fewer vehicles in the scene (and fewer lanes, shorter episode duration)
        - only check collision of controlled vehicles with others
    """
    @classmethod
    def default_config(cls) -> dict:
        cfg = super().default_config()
        cfg.update({
            "simulation_frequency": 5,
            "lanes_count": 3,
            "vehicles_count": 20,
            "duration": 30,  # [s]
            "ego_spacing": 1.5,
        })
        return cfg

    def _create_vehicles(self) -> None:
        super()._create_vehicles()
        # Disable collision check for uncontrolled vehicles
        for vehicle in self.road.vehicles:
            if vehicle not in self.controlled_vehicles:
                vehicle.check_collisions = False


class MyHighwayEnv(AbstractEnv):
    """
    A highway driving environment.

    The vehicle is driving on a straight highway with several lanes, and is rewarded for reaching a high speed,
    staying on the rightmost lanes and avoiding collisions.
    """

    @classmethod
    def default_config(cls) -> dict:
        config = super().default_config()
        config.update({
            "observation": {
                "type": "PotentialField", #PotentialField  OccupancyGrid
                "features": ['presence',"x", "y", "vx", "vy", "cos_h", "sin_h", 'potential_field'],
                "grid_size": [[-27.5, 27.5], [-27.5, 27.5]], #[-27.5, 27.5]
                "grid_step": [0.5, 0.5],
                "as_image": True,
                "align_to_vehicle_axes": False,
                "absolute": False
            },
            "action": {
                "type": "ContinuousAction",
                # "dynamical": True, # 먼저 학습 해 보고 잘되면 나중에 더하
            },
            "policy_frequency": 10,
            "decision_frequency": 10,
            "lanes_count": 4,
            "vehicles_count": 50,
            "controlled_vehicles": 1,
            "initial_lane_id": None,
            "duration": 40000000,  # [s]
            "ego_spacing": 2,
            "vehicles_density": 1,
            "collision_reward": -1,    # The reward received when colliding with a vehicle.
            "target_lane_reward": 0.8,  # The reward received when driving on the right-most lanes, linearly mapped
            "direction_reward":0.1,
            "high_speed_reward": 0.2,  # The reward received when driving at full speed, linearly mapped to zero for
                                       # lower speeds according to config["reward_speed_range"].
            "jerk_reward": 0,   # The reward received at each lane change action.
            "reward_speed_range": [20, 30],
            "offroad_terminal": True,
            "simulation_frequency" : 10
        })
        return config

    def _reset(self) -> None:
        self._create_road()
        self._create_vehicles()
        self._create_global_path()
        self.action_pre = np.zeros(self.action_space.shape, dtype=self.action_space.dtype)
        self.anvel = 0
        self.anvel_pre = 0
        self.trajectory = []
        self.target_lane = self.global_path[1][0]

        # x = 0
    def _create_global_path(self) -> None:
        """ create global path (reference path)"""
        LANE_LEN = self.controlled_vehicles[0].lane.length
        LANE_WID = self.controlled_vehicles[0].lane.width
        INIT_POS = np.append(self.controlled_vehicles[0].position, self.controlled_vehicles[0].lane_index[2])
        INTERVAL = 300.
        LANE_CNT = self.config['lanes_count']

        inter_x = np.arange(INIT_POS[0], LANE_LEN, INTERVAL)
        global_path = np.array([inter_x, np.zeros(inter_x.size)])
        self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        self.global_path = global_path

    def create_global_path(self, global_path, idx, lanes, lane_id):
        global_path[1][idx] = lane_id
        idx += 1
        if idx == global_path.size * 0.5:
            return global_path
        else : #0,1,2,3 -> lanes = 4
            if lane_id == 0:
                # global_path[1][idx] = 1
                lane_id = 1
            elif lane_id == lanes-1:
                # global_path[1][idx] = lane_id - 1
                lane_id = lane_id - 1
            else:
                # global_path[1][idx] = random.choice([lane_id-1, lane_id+1])
                # lane_id = global_path[1][idx]
                lane_id = random.choice([lane_id-1, lane_id+1])
            self.create_global_path(global_path, idx, lanes, lane_id)




    def _create_road(self) -> None:
        """Create a road composed of straight adjacent lanes."""
        self.road = Road(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                         np_random=self.np_random, record_history=self.config["show_trajectories"])

    def _create_vehicles(self) -> None:
        """Create some new random vehicles of a given type, and add them on the road."""
        other_vehicles_type = utils.class_from_path(self.config["other_vehicles_type"])
        other_per_controlled = near_split(self.config["vehicles_count"], num_bins=self.config["controlled_vehicles"])

        self.controlled_vehicles = []
        for others in other_per_controlled:
            vehicle = Vehicle.create_random(
                self.road,
                speed=25,
                lane_id=self.config["initial_lane_id"],
                spacing=self.config["ego_spacing"]
            )
            vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
            self.controlled_vehicles.append(vehicle)
            self.road.vehicles.append(vehicle)

            for _ in range(others):
                vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
                vehicle.randomize_behavior()
                self.road.vehicles.append(vehicle)

    def _reward(self, action: Action) -> float:
        """
        The reward is defined to foster driving at high speed, on the rightmost lanes, and to avoid collisions.
        :param action: the last action performed
        :return: the corresponding reward
        """
        neighbours = self.road.network.all_side_lanes(self.vehicle.lane_index)
        lane = self.vehicle.target_lane_index[2] if isinstance(self.vehicle, ControlledVehicle) \
            else self.vehicle.lane_index[2]
        width = self.road.network.graph['0']['1'][0].width
        self.anvel = action[1] - self.action_pre[1]
        scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [0, 1])
        jerk_v = abs(action[0] - self.action_pre[0])
        jerk_w = abs(self.anvel - self.anvel_pre)
        scaled_jerk_v = utils.lmap(jerk_v, [0, 10],[1, 0])
        scaled_jerk_w = utils.lmap(jerk_w, [0, np.pi],[1, 0])

        reward_c = self.config["collision_reward"] * self.vehicle.crashed
        reward_t = self.config["target_lane_reward"] * max(-1/(0.5*width)*np.abs(self.vehicle.position[1]-4*self.target_lane)+1,0)
        reward_s = self.config["high_speed_reward"] * np.clip(scaled_speed, 0, 1)
        # reward_d = self.config["direction_reward"] * (self.obs['decision'][0]-1)*np.sign(action[1])
        reward_jw = self.config["jerk_reward"] * (scaled_jerk_w)
        reward_jv = self.config["jerk_reward"] * scaled_jerk_v
        reward = reward_c + reward_t + reward_s+ reward_jv + reward_jw
        # print(
        #     "T : {:.2f}, St : {:.2f}, Sp : {:.2f}, Jv : {:.2f}, Jw : {:.2f}".format(reward_t/self.config["target_lane_reward"], reward_d/self.config["direction_reward"],
        #                                                                                             reward_s/self.config["high_speed_reward"], reward_jv/self.config["jerk_reward"],
        #                                                                                             reward_jw/self.config["jerk_reward"]))
        #
        # print("T : {:.2f}, St : {:.2f}, Sp : {:.2f}, Jv : {:.2f}, Jw : {:.2f}, Total : {:.2f}".format(reward_t, reward_d , reward_s, reward_jv , reward_jw, reward))
        # reward = utils.lmap(reward,
        #                   [self.config["collision_reward"],
        #                    self.config["high_speed_reward"] + self.config["right_lane_reward"]],
        #                   [0, 1])

        reward = -1 if not self.vehicle.on_road or self.vehicle.position[0]<self.trajectory[0][0] else reward


        self.action_pre = action
        self.anvel_pre = self.anvel

        return reward

    def _is_terminal(self) -> bool:
        """The episode is over if the ego vehicle crashed or the time is out."""

        return self.vehicle.crashed or \
            self.steps >= self.config["duration"] or \
            (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
            self.vehicle.position[0]<self.trajectory[0][0]

    def _cost(self, action: int) -> float:
        """The cost signal is the occurrence of collision."""
        return float(self.vehicle.crashed)


class HighwayEnv_Main(AbstractEnv):
    """
    A highway driving environment.

    The vehicle is driving on a straight highway with several lanes, and is rewarded for reaching a high speed,
    staying on the rightmost lanes and avoiding collisions.
    """

    @classmethod
    def default_config(cls) -> dict:
        config = super().default_config()
        config.update({
            "observation": {
                "type": "KinematicDecision2", #PotentialField  OccupancyGrid
            },
            "action": {
                "type": "ContinuousAction",
                'steering_range' : (-np.pi/30, np.pi/30),
                "dynamical": True, # 먼저 학습 해 보고 잘되면 나중에 더하
            },
            "policy_frequency": 10,
            "decision_frequency": 10,
            "lanes_count": 4,
            "vehicles_count": 50,
            "controlled_vehicles": 1,
            "initial_lane_id": None,
            "duration": 4000000,  # [s]
            "ego_spacing": 2,
            "vehicles_density": 1,
            "collision_reward": -5,    # The reward received when colliding with a vehicle.
            "target_lane_reward": 0.5,
            "direction_reward":0,
            "high_speed_reward": 0.5,  # The reward received when driving at full speed, linearly mapped to zero for
                                       # lower speeds according to config["reward_speed_range"].
            "jerk_w_reward": 0,
            "jerk_v_reward": 0,
            "reward_speed_range": [20, 30],
            "reward_speed_slope": [0.5, 1],
            "offroad_terminal": True,
            "simulation_frequency" : 10,
            "random_road" : False,
            "road_static" : False,
            'save_positions': False,
            'stop_speed': 10,
        })
        return config

    def _reset(self) -> None:
        if self.config['random_road']:
            self.config['lanes_count'] = random.randint(2,4)
        self.save_road = []
        self._create_road()
        self._create_vehicles()
        self._create_global_path()
        self.action_pre = np.zeros(self.action_space.shape, dtype=self.action_space.dtype)
        self.anvel = 0
        self.anvel_pre = 0
        self.trajectory = []
        self.target_lane = self.global_path[1][0]
        self.allowed_lane = [self.target_lane]
        self.decision_pre = 1
        self.decelerate_flag = False
        self.front_dist_stack = []

        # x = 0
    def _create_global_path(self) -> None:
        """ create global path (reference path)"""
        LANE_LEN = self.controlled_vehicles[0].lane.length
        LANE_WID = self.controlled_vehicles[0].lane.width
        INIT_POS = np.append(self.controlled_vehicles[0].position, self.controlled_vehicles[0].lane_index[2])
        INTERVAL = 800.#300.
        LANE_CNT = self.config['lanes_count']

        inter_x = np.arange(INIT_POS[0], LANE_LEN, INTERVAL)
        global_path = np.array([inter_x, np.zeros(inter_x.size)])
        self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        self.global_path = global_path

        if self.config['save_positions']:
            np.save('save_positions/save_global_'+str(self.config["lanes_count"])+'.npy', global_path)
            # write_wb = Workbook()
            # write_ws = write_wb.active
            # write_ws.append(global_path[0].tolist())
            # write_ws.append(global_path[1])
            # write_wb.save('save_positions/save_global_'+str(self.config["lanes_count"])+ '.xlsx')


        if self.config['road_static']:
            self.global_path = np.load('save_positions/save_global_'+str(self.config["lanes_count"])+ '.npy')
            ##
            # write_wb = Workbook()
            # write_ws = write_wb.active
            # write_ws.append(self.global_path[0].tolist())
            # write_ws.append(self.global_path[1].tolist())
            # write_wb.save('save_positions/save_global_'+str(self.config["lanes_count"])+ '.xlsx')



            # load_wb = load_workbook('save_positions/save_global_'+str(self.config["lanes_count"])+ '.xlsx', data_only = True)
            # load_ws = load_wb['Sheet1']
            # idx = 0
            # for row in load_ws.rows:
            #     self.global_path[idx] = row
            #     idx = idx+1


        # global_path = np.array([inter_x, np.zeros(inter_x.size)])
        # self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        # self.global_path = global_path

    def create_global_path(self, global_path, idx, lanes, lane_id):
        global_path[1][idx] = lane_id
        idx += 1
        if idx == global_path.size * 0.5:
            return global_path
        else : #0,1,2,3 -> lanes = 4
            if lane_id == 0:
                # global_path[1][idx] = 1
                lane_id = 1
            elif lane_id == lanes-1:
                # global_path[1][idx] = lane_id - 1
                lane_id = lane_id - 1
            else:
                # global_path[1][idx] = random.choice([lane_id-1, lane_id+1])
                # lane_id = global_path[1][idx]
                lane_id = random.choice([lane_id-1, lane_id+1])
            self.create_global_path(global_path, idx, lanes, lane_id)




    def _create_road(self) -> None:
        """Create a road composed of straight adjacent lanes."""
        if self.config["road_static"]:
            self.road = Road_Static(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])
        else:
            self.road = Road(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])

    def _create_vehicles(self) -> None:
        """Create some new random vehicles of a given type, and add them on the road."""
        other_vehicles_type = utils.class_from_path(self.config["other_vehicles_type"])
        other_per_controlled = near_split(self.config["vehicles_count"], num_bins=self.config["controlled_vehicles"])

        # if self.config["save_positions"]:
        #     road_static_data = []
        #     write_wb = Workbook()
        #     self.controlled_vehicles = []
        #
        #     for others in other_per_controlled:
        #         vehicle = Vehicle.create_random(
        #             self.road,
        #             speed=25,
        #             lane_id=self.config["initial_lane_id"],
        #             spacing=self.config["ego_spacing"]
        #         )
        #         vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
        #         self.controlled_vehicles.append(vehicle)
        #         self.road.vehicles.append(vehicle)
        #         road_static_data.append([vehicle.position[0], vehicle.position[1], vehicle.speed, vehicle.lane_index[0], \
        #                                  vehicle.lane_index[1], vehicle.lane_index[2]])
        #
        #         for _ in range(others):
        #             vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
        #             vehicle.randomize_behavior()
        #             self.road.vehicles.append(vehicle)
        #
        #             road_static_data.append(
        #                 [vehicle.position[0], vehicle.position[1], vehicle.speed, vehicle.lane_index[0], \
        #                  vehicle.lane_index[1], vehicle.lane_index[2]])
        #     col_names = ['pos x', 'pos y', 'speed', 'lane_idx1', 'lane_idx2', 'lane_idx3']
        #     write_ws = write_wb.active
        #     write_ws.append(col_names)
        #     for c in range(len(road_static_data)):
        #         write_ws.append(road_static_data[c])
        #     write_wb.save('save_positions/save_positions_'+str(self.config["lanes_count"])+ '.xlsx')




        # if not self.config['road_static']:
        self.controlled_vehicles = []
        for others in other_per_controlled:
            vehicle = Vehicle.create_random(
                self.road,
                speed=25,
                lane_id=self.config["initial_lane_id"],
                spacing=self.config["ego_spacing"]
            )
            vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
            self.controlled_vehicles.append(vehicle)
            self.road.vehicles.append(vehicle)

            for _ in range(others):
                vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
                vehicle.randomize_behavior()
                self.road.vehicles.append(vehicle)
        if self.config['save_positions']:
            np.save('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy', self.road.vehicles)
        if self.config['road_static']:
            self.road.vehicles = np.load('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy',
                                         allow_pickle=True)
            self.controlled_vehicles[0] = self.road.vehicles[0]

        # self.save_road = static_V(self.road.vehicles)
        # print(self.save_road)
    def _reward(self, action: Action) -> float:
        """
        The reward is defined to foster driving at high speed, on the rightmost lanes, and to avoid collisions.
        :param action: the last action performed
        :return: the corresponding reward
        """

        width = self.road.network.graph['0']['1'][0].width
        self.anvel = action[1] - self.action_pre[1]

        EGO_POS = np.append(self.controlled_vehicles[0].position,
                            [self.controlled_vehicles[0].speed, self.controlled_vehicles[0].lane_index[2]])



        front_dist_now = front_dist(EGO_POS, self.road.vehicles[1:])
        self.front_dist_stack.append(front_dist_now)

        if not self.decelerate_flag:
            if self.decision == 1 and front_dist_now <= EGO_POS[2]-5:
                self.decelerate_flag = True
        else:
            if self.decision  != 1 or front_dist_now > EGO_POS[2]:
                self.decelerate_flag = False
        # calculate reward
        if not self.decelerate_flag:
            if self.vehicle.speed > self.config["reward_speed_range"][1]:
                scaled_speed = 1
                # scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [3, 1])
            else:
                scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"],
                                          self.config["reward_speed_slope"])
        else:
            # print("DECELERATE FLAG ON")
            dspeed = front_speed(EGO_POS, self.road.vehicles[1:])
            if dspeed <= 0 :
                scaled_speed = utils.lmap(dspeed, [-1, 0], [0, 1])
            else:
                scaled_speed = utils.lmap(dspeed, [0, 1], [1, 0])
        ## 충돌 penalty에만 기대서 속도 제어를 하길 바랐지만 잘 안됨..... 감속 리워드가 필요함
        # if self.vehicle.speed < 30:
        #     scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [0, 1])
        # else:
        #     scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [2, 1])

        jerk_v = abs(action[0] - self.action_pre[0])
        jerk_w = abs(self.anvel - self.anvel_pre)
        scaled_jerk_v = utils.lmap(jerk_v, [0, 5],[1, 0])
        scaled_jerk_w = utils.lmap(jerk_w, [0, np.pi/30],[1, 0])

        reward_c = self.config["collision_reward"] * self.vehicle.crashed
        # reward_t = self.config["target_lane_reward"] * max(-1/(0.25*width)*np.abs(self.vehicle.position[1]-4*self.target_lane)+1,0)
        if self.decision  == 1:
            reward_t = self.config["target_lane_reward"] * max(-1/(0.25*width)*np.abs(self.vehicle.position[1]-4*self.target_lane)+1,0)
        else:
            # reward_t = self.config["target_lane_reward"] * max(-0.3/0.7*np.abs(self.vehicle.position[1]-self.lc_reward)+1,0)
            heading_reward = 0
            if self.decision  == 0 and self.controlled_vehicles[0].to_dict()['heading'] < 0 : # minus steering is positive
                heading_reward = 0.8
            elif self.decision  == 2 and self.controlled_vehicles[0].to_dict()['heading'] > 0 : # plus steering is positive
                heading_reward = 0.8
            reward_t = self.config["target_lane_reward"] * heading_reward

        reward_s = self.config["high_speed_reward"] * np.clip(scaled_speed, -2,2)
        reward_d = self.config["direction_reward"] * (self.decision -1)*np.sign(action[1])
        reward_jw = self.config["jerk_w_reward"] * np.clip(scaled_jerk_w,-1,1)
        reward_jv = self.config["jerk_v_reward"] * np.clip(scaled_jerk_v,-1,1)
        # print("REWARD : speed {:.2f}, lane {:.2f}".format(reward_s, reward_t))
        # print("steer : {:.2f}, ang vel : {:.2f}, jerk W : {:.2f}, ".format( action[1],utils.lmap(abs(self.anvel), [0, np.pi/10],[1, 0]) ,scaled_jerk_w))

        reward = reward_c + reward_t + reward_s+reward_d + reward_jv + reward_jw


        # reward = self.config["collision_reward"] if self.vehicle.crashed or \
        #          (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
        #          self.vehicle.position[0] < self.trajectory[0][0] or \
        #          self.controlled_vehicles[0].lane_index[2] not in self.allowed_lane or \
        #          self.controlled_vehicles[0].speed <= self.config['stop_speed'] else reward
        reward = self.config["collision_reward"] if self.vehicle.crashed or \
                 (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                 self.vehicle.position[0] < self.trajectory[0][0] or \
                 self.controlled_vehicles[0].speed <= self.config['stop_speed'] else reward

        self.action_pre = action
        self.anvel_pre = self.anvel
        self.decision_pre = self.decision
        return reward

    def _is_terminal(self) -> bool:
        """The episode is over if the ego vehicle crashed or the time is out."""

        isdone = self.vehicle.crashed or \
                 self.steps >= self.config["duration"] or \
                 (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                 self.vehicle.position[0] < self.trajectory[0][0] or \
                 self.controlled_vehicles[0].speed <= self.config['stop_speed']
                 # self.controlled_vehicles[0].lane_index[2] not in self.allowed_lane or \

        # if isdone:
        #     write_wb = Workbook()
        #     write_ws = write_wb.active
        #     write_ws.append(self.front_dist_stack)
        #     write_wb.save('/Users/cml/highway-env/scripts/highway_td3_decisionbuffer/success_test/driving style/CER4_3_2.xlsx')
        return isdone

    def _cost(self, action: int) -> float:
        """The cost signal is the occurrence of collision."""
        return float(self.vehicle.crashed)


class HighwayEnv_Long(AbstractEnv):
    """
    A highway driving environment.

    The vehicle is driving on a straight highway with several lanes, and is rewarded for reaching a high speed,
    staying on the rightmost lanes and avoiding collisions.
    """

    @classmethod
    def default_config(cls) -> dict:
        config = super().default_config()
        config.update({
            "observation": {
                "type": "KinematicDecision2", #PotentialField  OccupancyGrid
            },
            "action": {
                "type": "ContinuousAction",
                'steering_range' : (-np.pi/30, np.pi/30),
                "dynamical": True, # 먼저 학습 해 보고 잘되면 나중에 더하
            },
            "policy_frequency": 10,
            "decision_frequency": 10,
            "lanes_count": 4,
            "vehicles_count": 50,
            "controlled_vehicles": 1,
            "initial_lane_id": None,
            "duration": 4000000,  # [s]
            "ego_spacing": 2,
            "vehicles_density": 1,
            "collision_reward": -5,    # The reward received when colliding with a vehicle.
            "target_lane_reward": 0.5,
            "direction_reward":0,
            "high_speed_reward": 0.5,  # The reward received when driving at full speed, linearly mapped to zero for
                                       # lower speeds according to config["reward_speed_range"].
            "jerk_w_reward": 0,
            "jerk_v_reward": 0,
            "reward_speed_range": [20, 30],
            "reward_speed_slope": [0.5, 1],
            "offroad_terminal": True,
            "simulation_frequency" : 10,
            "random_road" : False,
            "road_static" : False,
            'save_positions': False,
            'stop_speed': 10,
        })
        return config

    def _reset(self) -> None:
        if self.config['random_road']:
            self.config['lanes_count'] = random.randint(2,4)
        self.save_road = []
        self._create_road()
        self._create_vehicles()
        self._create_global_path()
        self.action_pre = np.zeros(self.action_space.shape, dtype=self.action_space.dtype)
        self.anvel = 0
        self.anvel_pre = 0
        self.trajectory = []
        self.target_lane = self.global_path[1][0]
        self.allowed_lane = [self.target_lane]
        self.decision_pre = 1
        self.decelerate_flag = False


        # x = 0
    def _create_global_path(self) -> None:
        """ create global path (reference path)"""
        LANE_LEN = self.controlled_vehicles[0].lane.length
        LANE_WID = self.controlled_vehicles[0].lane.width
        INIT_POS = np.append(self.controlled_vehicles[0].position, self.controlled_vehicles[0].lane_index[2])
        INTERVAL = 800.
        LANE_CNT = self.config['lanes_count']

        inter_x = np.arange(INIT_POS[0], LANE_LEN, INTERVAL)
        global_path = np.array([inter_x, np.zeros(inter_x.size)])
        self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        self.global_path = global_path

        if self.config['save_positions']:
            np.save('save_positions/save_global_'+str(self.config["lanes_count"])+'.npy',global_path)
            # write_wb = Workbook()
            # write_ws = write_wb.active
            # write_ws.append(global_path[0].tolist())
            # write_ws.append(global_path[1])
            # write_wb.save('save_positions/save_global_'+str(self.config["lanes_count"])+ '.xlsx')


        if self.config['road_static']:
            self.global_path = np.load('save_positions/save_global_'+str(self.config["lanes_count"])+ '.npy')




            # load_wb = load_workbook('save_positions/save_global_'+str(self.config["lanes_count"])+ '.xlsx', data_only = True)
            # load_ws = load_wb['Sheet1']
            # idx = 0
            # for row in load_ws.rows:
            #     self.global_path[idx] = row
            #     idx = idx+1


        # global_path = np.array([inter_x, np.zeros(inter_x.size)])
        # self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        # self.global_path = global_path

    def create_global_path(self, global_path, idx, lanes, lane_id):
        global_path[1][idx] = lane_id
        idx += 1
        if idx == global_path.size * 0.5:
            return global_path
        else : #0,1,2,3 -> lanes = 4
            if lane_id == 0:
                # global_path[1][idx] = 1
                lane_id = 1
            elif lane_id == lanes-1:
                # global_path[1][idx] = lane_id - 1
                lane_id = lane_id - 1
            else:
                # global_path[1][idx] = random.choice([lane_id-1, lane_id+1])
                # lane_id = global_path[1][idx]
                lane_id = random.choice([lane_id-1, lane_id+1])
            self.create_global_path(global_path, idx, lanes, lane_id)




    def _create_road(self) -> None:
        """Create a road composed of straight adjacent lanes."""
        if self.config["road_static"]:
            self.road = Road_Static(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])
        else:
            self.road = Road(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])

    def _create_vehicles(self) -> None:
        """Create some new random vehicles of a given type, and add them on the road."""
        other_vehicles_type = utils.class_from_path(self.config["other_vehicles_type"])
        other_per_controlled = near_split(self.config["vehicles_count"], num_bins=self.config["controlled_vehicles"])

        # if self.config["save_positions"]:
        #     road_static_data = []
        #     write_wb = Workbook()
        #     self.controlled_vehicles = []
        #
        #     for others in other_per_controlled:
        #         vehicle = Vehicle.create_random(
        #             self.road,
        #             speed=25,
        #             lane_id=self.config["initial_lane_id"],
        #             spacing=self.config["ego_spacing"]
        #         )
        #         vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
        #         self.controlled_vehicles.append(vehicle)
        #         self.road.vehicles.append(vehicle)
        #         road_static_data.append([vehicle.position[0], vehicle.position[1], vehicle.speed, vehicle.lane_index[0], \
        #                                  vehicle.lane_index[1], vehicle.lane_index[2]])
        #
        #         for _ in range(others):
        #             vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
        #             vehicle.randomize_behavior()
        #             self.road.vehicles.append(vehicle)
        #
        #             road_static_data.append(
        #                 [vehicle.position[0], vehicle.position[1], vehicle.speed, vehicle.lane_index[0], \
        #                  vehicle.lane_index[1], vehicle.lane_index[2]])
        #     col_names = ['pos x', 'pos y', 'speed', 'lane_idx1', 'lane_idx2', 'lane_idx3']
        #     write_ws = write_wb.active
        #     write_ws.append(col_names)
        #     for c in range(len(road_static_data)):
        #         write_ws.append(road_static_data[c])
        #     write_wb.save('save_positions/save_positions_'+str(self.config["lanes_count"])+ '.xlsx')




        # if not self.config['road_static']:
        self.controlled_vehicles = []
        for others in other_per_controlled:
            vehicle = Vehicle.create_random(
                self.road,
                speed=25,
                lane_id=self.config["initial_lane_id"],
                spacing=self.config["ego_spacing"]
            )
            vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
            self.controlled_vehicles.append(vehicle)
            self.road.vehicles.append(vehicle)

            for _ in range(others):
                vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
                vehicle.randomize_behavior()
                self.road.vehicles.append(vehicle)
        if self.config['save_positions']:
            np.save('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy', self.road.vehicles)
        if self.config['road_static']:
            self.road.vehicles = np.load('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy',
                                         allow_pickle=True)
            self.controlled_vehicles[0] = self.road.vehicles[0]

        # self.save_road = static_V(self.road.vehicles)
        # print(self.save_road)
    def _reward(self, action: Action) -> float:
        """
        The reward is defined to foster driving at high speed, on the rightmost lanes, and to avoid collisions.
        :param action: the last action performed
        :return: the corresponding reward
        """

        width = self.road.network.graph['0']['1'][0].width
        self.anvel = action[1] - self.action_pre[1]

        EGO_POS = np.append(self.controlled_vehicles[0].position,
                            [self.controlled_vehicles[0].speed, self.controlled_vehicles[0].lane_index[2]])

        # scaled_dist = emap(front_dist(EGO_POS, self.road.vehicles[1:]))
        # scaled_dist = min(1, scaled_dist)
        # if self.vehicle.speed >= self.config["reward_speed_range"][0] and self.vehicle.speed <= \
        #         self.config["reward_speed_range"][1]:
        #     scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [0.5, 1])
        # elif self.vehicle.speed > self.config["reward_speed_range"][1]:
        #     if scaled_dist == 1:
        #         scaled_speed = 0
        #     else:
        #         scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [0.5, 1])
        # else:
        #     scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [-0.5, 0])
        ## change speed reward mode 05/02
        if not self.decelerate_flag:

            if self.decision == 1 and front_dist(EGO_POS, self.road.vehicles[1:]) <= EGO_POS[2]-5:
                self.decelerate_flag = True
        else:
            if self.decision  != 1 or front_dist(EGO_POS, self.road.vehicles[1:]) > EGO_POS[2]:
                self.decelerate_flag = False
        # calculate reward
        if not self.decelerate_flag:
            if self.vehicle.speed > self.config["reward_speed_range"][1]:
                scaled_speed = 1
                # scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [3, 1])
            else:
                scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"],
                                          self.config["reward_speed_slope"])
        else:
            # print("DECELERATE FLAG ON")
            dspeed = front_speed(EGO_POS, self.road.vehicles[1:])
            if dspeed <= 0 :
                scaled_speed = utils.lmap(dspeed, [-1, 0], [0, 1])
            else:
                scaled_speed = utils.lmap(dspeed, [0, 1], [1, 0])
        ## 충돌 penalty에만 기대서 속도 제어를 하길 바랐지만 잘 안됨..... 감속 리워드가 필요함
        # if self.vehicle.speed < 30:
        #     scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [0, 1])
        # else:
        #     scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [2, 1])

        jerk_v = abs(action[0] - self.action_pre[0])
        jerk_w = abs(self.anvel - self.anvel_pre)
        scaled_jerk_v = utils.lmap(jerk_v, [0, 5],[1, 0])
        scaled_jerk_w = utils.lmap(jerk_w, [0, np.pi/30],[1, 0])

        reward_c = self.config["collision_reward"] * self.vehicle.crashed
        # reward_t = self.config["target_lane_reward"] * max(-1/(0.25*width)*np.abs(self.vehicle.position[1]-4*self.target_lane)+1,0)
        if self.decision  == 1:
            reward_t = self.config["target_lane_reward"] * max(-1/(0.25*width)*np.abs(self.vehicle.position[1]-4*self.target_lane)+1,0)
        else:
            # reward_t = self.config["target_lane_reward"] * max(-0.3/0.7*np.abs(self.vehicle.position[1]-self.lc_reward)+1,0)
            heading_reward = 0
            if self.decision  == 0 and self.controlled_vehicles[0].to_dict()['heading'] < 0 : # minus steering is positive
                heading_reward = 0.8
            elif self.decision  == 2 and self.controlled_vehicles[0].to_dict()['heading'] > 0 : # plus steering is positive
                heading_reward = 0.8
            reward_t = self.config["target_lane_reward"] * heading_reward

        reward_s = self.config["high_speed_reward"] * np.clip(scaled_speed, -2,2)
        reward_d = self.config["direction_reward"] * (self.decision -1)*np.sign(action[1])
        reward_jw = self.config["jerk_w_reward"] * np.clip(scaled_jerk_w,-1,1)
        reward_jv = self.config["jerk_v_reward"] * np.clip(scaled_jerk_v,-1,1)
        # print("REWARD : speed {:.2f}, lane {:.2f}".format(reward_s, reward_t))
        # print("steer : {:.2f}, ang vel : {:.2f}, jerk W : {:.2f}, ".format( action[1],utils.lmap(abs(self.anvel), [0, np.pi/10],[1, 0]) ,scaled_jerk_w))

        reward = reward_c + reward_t + reward_s+reward_d + reward_jv + reward_jw


        # reward = self.config["collision_reward"] if self.vehicle.crashed or \
        #          (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
        #          self.vehicle.position[0] < self.trajectory[0][0] or \
        #          self.controlled_vehicles[0].lane_index[2] not in self.allowed_lane or \
        #          self.controlled_vehicles[0].speed <= self.config['stop_speed'] else reward
        reward = self.config["collision_reward"] if self.vehicle.crashed or \
                 (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                 self.vehicle.position[0] < self.trajectory[0][0] or \
                 self.controlled_vehicles[0].speed <= self.config['stop_speed'] else reward

        self.action_pre = action
        self.anvel_pre = self.anvel
        self.decision_pre = self.decision
        return reward

    def _is_terminal(self) -> bool:
        """The episode is over if the ego vehicle crashed or the time is out."""

        isdone = self.vehicle.crashed or \
                 self.steps >= self.config["duration"] or \
                 (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                 self.vehicle.position[0] < self.trajectory[0][0] or \
                 self.controlled_vehicles[0].speed <= self.config['stop_speed']
                 # self.controlled_vehicles[0].lane_index[2] not in self.allowed_lane or \


        return isdone

    def _cost(self, action: int) -> float:
        """The cost signal is the occurrence of collision."""
        return float(self.vehicle.crashed)


class HighwayEnv_Simple(AbstractEnv):
    """
    A highway driving environment.

    The vehicle is driving on a straight highway with several lanes, and is rewarded for reaching a high speed,
    staying on the rightmost lanes and avoiding collisions.
    """

    @classmethod
    def default_config(cls) -> dict:
        config = super().default_config()
        config.update({
            "observation": {
                "type": "KinematicDecision2", #PotentialField  OccupancyGrid
            },
            "action": {
                "type": "ContinuousAction",
                'steering_range' : (-np.pi/30, np.pi/30),
                "dynamical": True, # 먼저 학습 해 보고 잘되면 나중에 더하
            },
            "policy_frequency": 10,
            "decision_frequency": 10,
            "lanes_count": 4,
            "vehicles_count": 50,
            "controlled_vehicles": 1,
            "initial_lane_id": None,
            "duration": 4000000,  # [s]
            "ego_spacing": 2,
            "vehicles_density": 1,
            "collision_reward": -5,    # The reward received when colliding with a vehicle.
            "target_lane_reward": 0.5,
            "direction_reward":0,
            "high_speed_reward": 0.5,  # The reward received when driving at full speed, linearly mapped to zero for
                                       # lower speeds according to config["reward_speed_range"].
            "jerk_w_reward": 0,
            "jerk_v_reward": 0,
            "reward_speed_range": [20, 30],
            "reward_speed_slope": [0.5, 1],
            "offroad_terminal": True,
            "simulation_frequency" : 10,
            "random_road" : False,
            "road_static" : False,
            'save_positions': False,
            'stop_speed': 10,
        })
        return config

    def _reset(self) -> None:
        if self.config['random_road']:
            self.config['lanes_count'] = random.randint(2,4)
        self.save_road = []
        self._create_road()
        self._create_vehicles()
        self._create_global_path()
        self.action_pre = np.zeros(self.action_space.shape, dtype=self.action_space.dtype)
        self.anvel = 0
        self.anvel_pre = 0
        self.trajectory = []
        self.target_lane = self.global_path[1][0]
        self.allowed_lane = [self.target_lane]
        self.decision_pre = 1
        self.decelerate_flag = False


        # x = 0
    def _create_global_path(self) -> None:
        """ create global path (reference path)"""
        LANE_LEN = self.controlled_vehicles[0].lane.length
        LANE_WID = self.controlled_vehicles[0].lane.width
        INIT_POS = np.append(self.controlled_vehicles[0].position, self.controlled_vehicles[0].lane_index[2])
        INTERVAL = 300.
        LANE_CNT = self.config['lanes_count']

        inter_x = np.arange(INIT_POS[0], LANE_LEN, INTERVAL)
        global_path = np.array([inter_x, np.zeros(inter_x.size)])
        self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        self.global_path = global_path

        if self.config['save_positions']:
            np.save('save_positions/save_global_'+str(self.config["lanes_count"])+'.npy',global_path)
            # write_wb = Workbook()
            # write_ws = write_wb.active
            # write_ws.append(global_path[0].tolist())
            # write_ws.append(global_path[1])
            # write_wb.save('save_positions/save_global_'+str(self.config["lanes_count"])+ '.xlsx')


        if self.config['road_static']:
            self.global_path = np.load('save_positions/save_global_'+str(self.config["lanes_count"])+ '.npy')




            # load_wb = load_workbook('save_positions/save_global_'+str(self.config["lanes_count"])+ '.xlsx', data_only = True)
            # load_ws = load_wb['Sheet1']
            # idx = 0
            # for row in load_ws.rows:
            #     self.global_path[idx] = row
            #     idx = idx+1


        # global_path = np.array([inter_x, np.zeros(inter_x.size)])
        # self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        # self.global_path = global_path

    def create_global_path(self, global_path, idx, lanes, lane_id):
        global_path[1][idx] = lane_id
        idx += 1
        if idx == global_path.size * 0.5:
            return global_path
        else : #0,1,2,3 -> lanes = 4
            if lane_id == 0:
                # global_path[1][idx] = 1
                lane_id = 1
            elif lane_id == lanes-1:
                # global_path[1][idx] = lane_id - 1
                lane_id = lane_id - 1
            else:
                # global_path[1][idx] = random.choice([lane_id-1, lane_id+1])
                # lane_id = global_path[1][idx]
                lane_id = random.choice([lane_id-1, lane_id+1])
            self.create_global_path(global_path, idx, lanes, lane_id)




    def _create_road(self) -> None:
        """Create a road composed of straight adjacent lanes."""
        if self.config["road_static"]:
            self.road = Road_Static(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])
        else:
            self.road = Road(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])

    def _create_vehicles(self) -> None:
        """Create some new random vehicles of a given type, and add them on the road."""
        other_vehicles_type = utils.class_from_path(self.config["other_vehicles_type"])
        other_per_controlled = near_split(self.config["vehicles_count"], num_bins=self.config["controlled_vehicles"])

        # if self.config["save_positions"]:
        #     road_static_data = []
        #     write_wb = Workbook()
        #     self.controlled_vehicles = []
        #
        #     for others in other_per_controlled:
        #         vehicle = Vehicle.create_random(
        #             self.road,
        #             speed=25,
        #             lane_id=self.config["initial_lane_id"],
        #             spacing=self.config["ego_spacing"]
        #         )
        #         vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
        #         self.controlled_vehicles.append(vehicle)
        #         self.road.vehicles.append(vehicle)
        #         road_static_data.append([vehicle.position[0], vehicle.position[1], vehicle.speed, vehicle.lane_index[0], \
        #                                  vehicle.lane_index[1], vehicle.lane_index[2]])
        #
        #         for _ in range(others):
        #             vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
        #             vehicle.randomize_behavior()
        #             self.road.vehicles.append(vehicle)
        #
        #             road_static_data.append(
        #                 [vehicle.position[0], vehicle.position[1], vehicle.speed, vehicle.lane_index[0], \
        #                  vehicle.lane_index[1], vehicle.lane_index[2]])
        #     col_names = ['pos x', 'pos y', 'speed', 'lane_idx1', 'lane_idx2', 'lane_idx3']
        #     write_ws = write_wb.active
        #     write_ws.append(col_names)
        #     for c in range(len(road_static_data)):
        #         write_ws.append(road_static_data[c])
        #     write_wb.save('save_positions/save_positions_'+str(self.config["lanes_count"])+ '.xlsx')




        # if not self.config['road_static']:
        self.controlled_vehicles = []
        for others in other_per_controlled:
            vehicle = Vehicle.create_random(
                self.road,
                speed=25,
                lane_id=self.config["initial_lane_id"],
                spacing=self.config["ego_spacing"]
            )
            vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
            self.controlled_vehicles.append(vehicle)
            self.road.vehicles.append(vehicle)

            for _ in range(others):
                vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
                vehicle.randomize_behavior()
                self.road.vehicles.append(vehicle)
        if self.config['save_positions']:
            np.save('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy', self.road.vehicles)
        if self.config['road_static']:
            self.road.vehicles = np.load('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy',
                                         allow_pickle=True)
            self.controlled_vehicles[0] = self.road.vehicles[0]

        # self.save_road = static_V(self.road.vehicles)
        # print(self.save_road)
    def _reward(self, action: Action) -> float:
        """
        The reward is defined to foster driving at high speed, on the rightmost lanes, and to avoid collisions.
        :param action: the last action performed
        :return: the corresponding reward
        """

        width = self.road.network.graph['0']['1'][0].width
        self.anvel = action[1] - self.action_pre[1]

        EGO_POS = np.append(self.controlled_vehicles[0].position,
                            [self.controlled_vehicles[0].speed, self.controlled_vehicles[0].lane_index[2]])

        ## decelerate reward 제거
        # if not self.decelerate_flag:
        #
        #     if self.decision == 1 and front_dist(EGO_POS, self.road.vehicles[1:]) <= EGO_POS[2]-5:
        #         self.decelerate_flag = True
        # else:
        #     if self.decision  != 1 or front_dist(EGO_POS, self.road.vehicles[1:]) > EGO_POS[2]:
        #         self.decelerate_flag = False
        # # calculate reward
        # if not self.decelerate_flag:
        #     if self.vehicle.speed > self.config["reward_speed_range"][1]:
        #         scaled_speed = 1
        #         # scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], [3, 1])
        #     else:
        #         scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"],
        #                                   self.config["reward_speed_slope"])
        # else:
        #     # print("DECELERATE FLAG ON")
        #     dspeed = front_speed(EGO_POS, self.road.vehicles[1:])
        #     if dspeed <= 0 :
        #         scaled_speed = utils.lmap(dspeed, [-1, 0], [0, 1])
        #     else:
        #         scaled_speed = utils.lmap(dspeed, [0, 1], [1, 0])

        scaled_speed = utils.lmap(self.vehicle.speed, self.config["reward_speed_range"], self.config["reward_speed_slope"])

        jerk_v = abs(action[0] - self.action_pre[0])
        jerk_w = abs(self.anvel - self.anvel_pre)
        scaled_jerk_v = utils.lmap(jerk_v, [0, 5],[1, 0])
        scaled_jerk_w = utils.lmap(jerk_w, [0, np.pi/30],[1, 0])

        reward_c = self.config["collision_reward"] * self.vehicle.crashed
        # reward_t = self.config["target_lane_reward"] * max(-1/(0.25*width)*np.abs(self.vehicle.position[1]-4*self.target_lane)+1,0)

        if self.decision  == 1:
            if self.decision_pre != 1: #lane change 가 끝남!
                lat_pos = 1
            else:
                lat_pos = 1 if np.abs(self.vehicle.position[1]-4*self.target_lane) < 0.05*width else 0
        else:
            lat_pos = 0
        # else:
        #     # reward_t = self.config["target_lane_reward"] * max(-0.3/0.7*np.abs(self.vehicle.position[1]-self.lc_reward)+1,0)
        #     heading_reward = 0
        #     if self.decision  == 0 and self.controlled_vehicles[0].to_dict()['heading'] < 0 : # minus steering is positive
        #         heading_reward = 0.8
        #     elif self.decision  == 2 and self.controlled_vehicles[0].to_dict()['heading'] > 0 : # plus steering is positive
        #         heading_reward = 0.8
        #     reward_t = self.config["target_lane_reward"] * heading_reward

        reward_s = self.config["high_speed_reward"] * np.clip(scaled_speed, -2,2)*0.5
        # reward_d = self.config["direction_reward"] * (self.decision -1)*np.sign(action[1])
        reward_jw = self.config["jerk_w_reward"] * np.clip(scaled_jerk_w,-1,1)
        reward_jv = self.config["jerk_v_reward"] * np.clip(scaled_jerk_v,-1,1)
        reward_t = self.config["target_lane_reward"] * lat_pos
        # print("REWARD : speed {:.2f}, lane {:.2f}".format(reward_s, reward_t))
        # print("steer : {:.2f}, ang vel : {:.2f}, jerk W : {:.2f}, ".format( action[1],utils.lmap(abs(self.anvel), [0, np.pi/10],[1, 0]) ,scaled_jerk_w))

        reward = reward_c + reward_t + reward_s+ reward_jv + reward_jw


        # reward = self.config["collision_reward"] if self.vehicle.crashed or \
        #          (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
        #          self.vehicle.position[0] < self.trajectory[0][0] or \
        #          self.controlled_vehicles[0].lane_index[2] not in self.allowed_lane or \
        #          self.controlled_vehicles[0].speed <= self.config['stop_speed'] else reward
        reward = self.config["collision_reward"] if self.vehicle.crashed or \
                 (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                 self.vehicle.position[0] < self.trajectory[0][0] or \
                 self.controlled_vehicles[0].speed <= self.config['stop_speed'] else reward

        self.action_pre = action
        self.anvel_pre = self.anvel
        self.decision_pre = self.decision
        return reward

    def _is_terminal(self) -> bool:
        """The episode is over if the ego vehicle crashed or the time is out."""

        isdone = self.vehicle.crashed or \
                 self.steps >= self.config["duration"] or \
                 (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                 self.vehicle.position[0] < self.trajectory[0][0] or \
                 self.controlled_vehicles[0].speed <= self.config['stop_speed']
                 # self.controlled_vehicles[0].lane_index[2] not in self.allowed_lane or \


        return isdone

    def _cost(self, action: int) -> float:
        """The cost signal is the occurrence of collision."""
        return float(self.vehicle.crashed)

class HighwayEnv_Reference_Reward(AbstractEnv):
    """
    A highway driving environment.

    The vehicle is driving on a straight highway with several lanes, and is rewarded for reaching a high speed,
    staying on the rightmost lanes and avoiding collisions.
    """

    @classmethod
    def default_config(cls) -> dict:
        config = super().default_config()
        config.update({
            "observation": {
                "type": "KinematicDecision2", #PotentialField  OccupancyGrid
            },
            "action": {
                "type": "ContinuousAction",
                'steering_range' : (-np.pi/10, np.pi/10),

                # "dynamical": True, # 먼저 학습 해 보고 잘되면 나중에 더하
            },
            "policy_frequency": 10,
            "decision_frequency": 10,
            "lanes_count": 4,
            "vehicles_count": 50,
            "controlled_vehicles": 1,
            "initial_lane_id": None,
            "duration": 4000000,  # [s]
            "ego_spacing": 2,
            "vehicles_density": 1,
            "collision_reward": -5,    # The reward received when colliding with a vehicle.
            "target_lane_reward": 0.5,
            "direction_reward":0,
            "high_speed_reward": 0.5,  # The reward received when driving at full speed, linearly mapped to zero for
                                       # lower speeds according to config["reward_speed_range"].
            "jerk_w_reward": 0,
            "jerk_v_reward": 0,
            "reward_speed_range": [20, 30],
            "reward_speed_slope": [0.5, 1],
            "offroad_terminal": True,
            "simulation_frequency" : 10,
            "random_road" : True,
            "road_static" : False,
            'save_positions': False,
            'stop_speed': 15,
            'mpc_path' : None,
            'train' : True,
        })
        return config

    def _reset(self) -> None:
        if self.config['random_road']:
            self.config['lanes_count'] = random.randint(2,4)

        self.save_road = []
        self._create_road()
        self._create_vehicles()
        self._create_global_path()
        self.action_pre = np.zeros(self.action_space.shape, dtype=self.action_space.dtype)
        self.anvel = 0
        self.anvel_pre = 0
        self.trajectory = []
        self.target_lane = self.global_path[1][0]
        self.allowed_lane = [self.target_lane]
        self.decision_pre = 1
        self.decelerate_flag = False
        self.mpc_env = MPCObservation_Simple(self)
        self.mpc_action = [0, 0]
        global mpc_mpdel, mpc
        mpc_model = Model()
        mpc = MPC(mpc_model)
        mpc.set_initial_guess()

        if self.config['mpc_path'] is not None:
            self.ref_path = np.load(self.config['mpc_path'] + '/MPCpath_4.npy')

        # self.mpc_model = Model()
        # self.mpc = MPC(self.mpc_model)
        # self.mpc_env = MPCObservation_Simple(self)
        # self.mpc_obs = self.mpc_env.observe()
        # self.mpc.set_initial_guess()
        #
        # self.u0 = 0

        # x = 0
    def _create_global_path(self) -> None:
        """ create global path (reference path)"""
        LANE_LEN = self.controlled_vehicles[0].lane.length
        LANE_WID = self.controlled_vehicles[0].lane.width
        INIT_POS = np.append(self.controlled_vehicles[0].position, self.controlled_vehicles[0].lane_index[2])
        INTERVAL = 800.
        LANE_CNT = self.config['lanes_count']

        inter_x = np.arange(INIT_POS[0], LANE_LEN, INTERVAL)
        global_path = np.array([inter_x, np.zeros(inter_x.size)])
        self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        self.global_path = global_path

        if self.config['save_positions']:
            np.save('save_positions/save_global_'+str(self.config["lanes_count"])+'.npy',global_path)
            # write_wb = Workbook()
            # write_ws = write_wb.active
            # write_ws.append(global_path[0].tolist())
            # write_ws.append(global_path[1])
            # write_wb.save('save_positions/save_global_'+str(self.config["lanes_count"])+ '.xlsx')


        if self.config['road_static']:
            self.global_path = np.load('save_positions/save_global_'+str(self.config["lanes_count"])+ '.npy')




            # load_wb = load_workbook('save_positions/save_global_'+str(self.config["lanes_count"])+ '.xlsx', data_only = True)
            # load_ws = load_wb['Sheet1']
            # idx = 0
            # for row in load_ws.rows:
            #     self.global_path[idx] = row
            #     idx = idx+1


        # global_path = np.array([inter_x, np.zeros(inter_x.size)])
        # self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        # self.global_path = global_path

    def create_global_path(self, global_path, idx, lanes, lane_id):
        global_path[1][idx] = lane_id
        idx += 1
        if idx == global_path.size * 0.5:
            return global_path
        else : #0,1,2,3 -> lanes = 4
            if lane_id == 0:
                # global_path[1][idx] = 1
                lane_id = 1
            elif lane_id == lanes-1:
                # global_path[1][idx] = lane_id - 1
                lane_id = lane_id - 1
            else:
                # global_path[1][idx] = random.choice([lane_id-1, lane_id+1])
                # lane_id = global_path[1][idx]
                lane_id = random.choice([lane_id-1, lane_id+1])
            self.create_global_path(global_path, idx, lanes, lane_id)

    def _create_road(self) -> None:
        """Create a road composed of straight adjacent lanes."""
        if self.config["road_static"]:
            self.road = Road_Static(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])
        else:
            self.road = Road(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])

    def _create_vehicles(self) -> None:
        """Create some new random vehicles of a given type, and add them on the road."""
        other_vehicles_type = utils.class_from_path(self.config["other_vehicles_type"])
        other_per_controlled = near_split(self.config["vehicles_count"], num_bins=self.config["controlled_vehicles"])

        # if self.config["save_positions"]:
        #     road_static_data = []
        #     write_wb = Workbook()
        #     self.controlled_vehicles = []
        #
        #     for others in other_per_controlled:
        #         vehicle = Vehicle.create_random(
        #             self.road,
        #             speed=25,
        #             lane_id=self.config["initial_lane_id"],
        #             spacing=self.config["ego_spacing"]
        #         )
        #         vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
        #         self.controlled_vehicles.append(vehicle)
        #         self.road.vehicles.append(vehicle)
        #         road_static_data.append([vehicle.position[0], vehicle.position[1], vehicle.speed, vehicle.lane_index[0], \
        #                                  vehicle.lane_index[1], vehicle.lane_index[2]])
        #
        #         for _ in range(others):
        #             vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
        #             vehicle.randomize_behavior()
        #             self.road.vehicles.append(vehicle)
        #
        #             road_static_data.append(
        #                 [vehicle.position[0], vehicle.position[1], vehicle.speed, vehicle.lane_index[0], \
        #                  vehicle.lane_index[1], vehicle.lane_index[2]])
        #     col_names = ['pos x', 'pos y', 'speed', 'lane_idx1', 'lane_idx2', 'lane_idx3']
        #     write_ws = write_wb.active
        #     write_ws.append(col_names)
        #     for c in range(len(road_static_data)):
        #         write_ws.append(road_static_data[c])
        #     write_wb.save('save_positions/save_positions_'+str(self.config["lanes_count"])+ '.xlsx')




        # if not self.config['road_static']:
        self.controlled_vehicles = []
        for others in other_per_controlled:
            vehicle = Vehicle.create_random(
                self.road,
                speed=25,
                lane_id=self.config["initial_lane_id"],
                spacing=self.config["ego_spacing"]
            )
            vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
            self.controlled_vehicles.append(vehicle)
            self.road.vehicles.append(vehicle)

            for _ in range(others):
                vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
                vehicle.randomize_behavior()
                self.road.vehicles.append(vehicle)
        if self.config['save_positions']:
            np.save('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy', self.road.vehicles)
        if self.config['road_static']:
            self.road.vehicles = np.load('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy',
                                         allow_pickle=True)
            self.controlled_vehicles[0] = self.road.vehicles[0]

        # self.save_road = static_V(self.road.vehicles)
        # print(self.save_road)

    def mpc_observation(self):
        mpc_env = copy.deepcopy(self)
        self.mpc_action = mpc.make_step(self.mpc_obs)

        if not mpc.data.success[-1][0]:
            # print(list(round(i[0],2) for i in self.mpc_obs))
            self.mpc_action = self.mpc_action_pre
        #u0은 [-1~1] 사이의 값, self.mpc_action, self.controlled_vehicles.action 은 action range사이의 값임.
        self.mpc_action = self.mpc_action.reshape(1, 2)[0]
        # u0 = copy.deepcopy(self.mpc_action)
        # u0[0] = utils.lmap(u0[0], self.config['action']['acceleration_range'], [-1, 1])
        # u0[1] = utils.lmap(u0[1], self.config['action']['steering_range'], [-1, 1])
        #
        # mpc_env._simulate(u0)
        # # mpc_obs = MPCObservation_Simple(mpc_env).observe()
        # mpc_obs_env = MPCObservation_Simple(mpc_env)
        # mpc_obs_env.decision = self.observation_type.decision
        # mpc_obs_env.target_lane = self.observation_type.target_lane
        # mpc_obs = mpc_obs_env.observe()

        self.mpc_action_pre = self.mpc_action
        # return mpc_obs
        return None

    # def mpc_reward(self):
    #     if self.config['mpc_path'] is None:
    #         reward_a = -math.sqrt((self.mpc_action[1] - self.controlled_vehicles[0].action['steering']) ** 2) / (
    #                 np.pi / 10 * 2) + 1
    #         reward_s = -math.sqrt((self.mpc_action[0] - self.controlled_vehicles[0].action['acceleration']) ** 2) / (5 * 2) + 1
    #         reward = reward_s + reward_a
    #     else:
    #         #['pos x', 'pos y', 'steering', 'acc', 'speed', 'solver', 'target y', 'heading','distance','input1','input2']
    #         reward_a = -math.sqrt((self.ref_path[self.steps][3] - self.controlled_vehicles[0].action['acceleration']) ** 2) / (
    #                 np.pi / 10 * 2) + 1
    #         reward_s = -math.sqrt((self.ref_path[self.steps][3] - self.controlled_vehicles[0].action['steering']) ** 2) / (
    #                     5 * 2) + 1
    #
    #
    #     return reward

    def _reward(self, action: Action) -> float:
        """
        The reward is defined to foster driving at high speed, on the rightmost lanes, and to avoid collisions.
        :param action: the last action performed
        :return: the corresponding reward
        """
        if self.config['train']:
            EGO_POS = np.append(self.controlled_vehicles[0].position,
                                [self.controlled_vehicles[0].speed, self.controlled_vehicles[0].lane_index[2]])

            if self.config['mpc_path'] is None:
                parm_t, parm_a, parm_s = 0.4, 0.3, 0.3
                self.mpc_env.decision = self.observation_type.decision
                self.mpc_env.target_lane = self.observation_type.target_lane
                self.mpc_obs = self.mpc_env.observe()

                reward_t =max(-1 / (0.25 * 4) * np.abs(self.vehicle.position[1] - 4 * self.target_lane) + 1, 0)
                reward_a = -math.sqrt((self.mpc_action[1] - self.controlled_vehicles[0].action['steering']) ** 2) / (
                        np.pi / 10 * 2) + 1
                reward_s = 1-math.sqrt((self.mpc_action[0] - self.controlled_vehicles[0].action['acceleration'])**2) / (5*2)
                reward = parm_t * reward_t + parm_s*reward_s + parm_a*reward_a
                self.mpc_obs_next = self.mpc_observation()

            else:
                #['pos x', 'pos y', 'steering', 'acc', 'speed', 'solver', 'target y', 'heading','distance','input1','input2']
                parm_a, parm_s, parm_x, parm_y = 0.25, 0.25, 0.25, 0.25
                reward_a = np.clip(-math.sqrt((self.ref_path[self.steps][3] - self.controlled_vehicles[0].action['acceleration']) ** 2) / (
                        np.pi / 10 * 2) + 1, 0, 1)
                reward_s = np.clip(-math.sqrt((self.ref_path[self.steps][2] - self.controlled_vehicles[0].action['steering']) ** 2) / (
                            5 * 2) + 1, 0, 1)
                reward_x = np.clip(-math.sqrt(
                    (self.ref_path[self.steps][0] - self.controlled_vehicles[0].position[0]) ** 2) / (
                                   10) + 1, -1,1)
                reward_y = np.clip(-math.sqrt(
                    (self.ref_path[self.steps][1] - self.controlled_vehicles[0].position[1]) ** 2) / (
                                   4) + 1, -1,1)
                reward = parm_a*reward_a + parm_s*reward_s + parm_x*reward_x + parm_y*reward_y
            # reward = self.mpc_reward()

            reward = self.config["collision_reward"] if self.vehicle.crashed or \
                     (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                     self.vehicle.position[0] < self.trajectory[0][0] or \
                     self.controlled_vehicles[0].speed <= self.config['stop_speed'] else reward

        else:
            reward = 0





        return reward

    def _is_terminal(self) -> bool:
        """The episode is over if the ego vehicle crashed or the time is out."""

        isdone = self.vehicle.crashed or \
                 self.steps >= self.config["duration"] or \
                 (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                 self.vehicle.position[0] < self.trajectory[0][0] or \
                 self.controlled_vehicles[0].speed <= self.config['stop_speed']
                 # self.controlled_vehicles[0].lane_index[2] not in self.allowed_lane or \


        return isdone

    def _cost(self, action: int) -> float:
        """The cost signal is the occurrence of collision."""
        return float(self.vehicle.crashed)

class HighwayEnv_MPC_Buffer(AbstractEnv):
    """
    A highway driving environment.

    The vehicle is driving on a straight highway with several lanes, and is rewarded for reaching a high speed,
    staying on the rightmost lanes and avoiding collisions.
    """

    @classmethod
    def default_config(cls) -> dict:
        config = super().default_config()
        config.update({
            "observation": {
                "type": "KinematicDecision2", #PotentialField  OccupancyGrid
            },
            "action": {
                "type": "ContinuousAction",
                'steering_range' : (-np.pi/10, np.pi/10),

                # "dynamical": True, # 먼저 학습 해 보고 잘되면 나중에 더하
            },
            "policy_frequency": 10,
            "decision_frequency": 10,
            "lanes_count": 4,
            "vehicles_count": 50,
            "controlled_vehicles": 1,
            "initial_lane_id": None,
            "duration": 4000000,  # [s]
            "ego_spacing": 2,
            "vehicles_density": 1,
            "collision_reward": -5,    # The reward received when colliding with a vehicle.
            "target_lane_reward": 0.5,
            "direction_reward":0,
            "high_speed_reward": 0.5,  # The reward received when driving at full speed, linearly mapped to zero for
                                       # lower speeds according to config["reward_speed_range"].
            "jerk_w_reward": 0,
            "jerk_v_reward": 0,
            "reward_speed_range": [20, 30],
            "reward_speed_slope": [0.5, 1],
            "offroad_terminal": True,
            "simulation_frequency" : 10,
            "random_road" : True,
            "road_static" : False,
            'save_positions': False,
            'stop_speed': 15,
            'mpc_path' : None,
            'train' : True,
        })
        return config

    def _reset(self) -> None:
        if self.config['random_road']:
            self.config['lanes_count'] = random.randint(2,4)

        self.save_road = []
        self._create_road()
        self._create_vehicles()
        self._create_global_path()
        self.action_pre = np.zeros(self.action_space.shape, dtype=self.action_space.dtype)
        self.anvel = 0
        self.anvel_pre = 0
        self.trajectory = []
        self.target_lane = self.global_path[1][0]
        self.allowed_lane = [self.target_lane]
        self.decision_pre = 1
        self.steer_pre = 0
        self.acc_pre = 0
        self.decelerate_flag = False
        self.mpc_env = MPCObservation(self)
        self.mpc_action = [0, 0]
        global mpc_mpdel, mpc
        mpc_model = Model()
        mpc = MPC(mpc_model)
        mpc.set_initial_guess()

        if self.config['mpc_path'] is not None:
            self.ref_path = np.load(self.config['mpc_path'] + '/MPCpath_4.npy')

    def _create_global_path(self) -> None:
        """ create global path (reference path)"""
        LANE_LEN = self.controlled_vehicles[0].lane.length
        LANE_WID = self.controlled_vehicles[0].lane.width
        INIT_POS = np.append(self.controlled_vehicles[0].position, self.controlled_vehicles[0].lane_index[2])
        INTERVAL = 800.
        LANE_CNT = self.config['lanes_count']

        inter_x = np.arange(INIT_POS[0], LANE_LEN, INTERVAL)
        global_path = np.array([inter_x, np.zeros(inter_x.size)])
        self.create_global_path(global_path, 0, LANE_CNT, INIT_POS[2])
        self.global_path = global_path

        if self.config['save_positions']:
            np.save('save_positions/save_global_'+str(self.config["lanes_count"])+'.npy',global_path)

        if self.config['road_static']:
            self.global_path = np.load('save_positions/save_global_'+str(self.config["lanes_count"])+ '.npy')

    def create_global_path(self, global_path, idx, lanes, lane_id):
        global_path[1][idx] = lane_id
        idx += 1
        if idx == global_path.size * 0.5:
            return global_path
        else : #0,1,2,3 -> lanes = 4
            if lane_id == 0:
                # global_path[1][idx] = 1
                lane_id = 1
            elif lane_id == lanes-1:
                # global_path[1][idx] = lane_id - 1
                lane_id = lane_id - 1
            else:
                # global_path[1][idx] = random.choice([lane_id-1, lane_id+1])
                # lane_id = global_path[1][idx]
                lane_id = random.choice([lane_id-1, lane_id+1])
            self.create_global_path(global_path, idx, lanes, lane_id)

    def _create_road(self) -> None:
        """Create a road composed of straight adjacent lanes."""
        if self.config["road_static"]:
            self.road = Road_Static(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])
        else:
            self.road = Road(network=RoadNetwork.straight_road_network(self.config["lanes_count"], speed_limit=30),
                             np_random=self.np_random, record_history=self.config["show_trajectories"])

    def _create_vehicles(self) -> None:
        """Create some new random vehicles of a given type, and add them on the road."""
        other_vehicles_type = utils.class_from_path(self.config["other_vehicles_type"])
        other_per_controlled = near_split(self.config["vehicles_count"], num_bins=self.config["controlled_vehicles"])

        self.controlled_vehicles = []
        for others in other_per_controlled:
            vehicle = Vehicle.create_random(
                self.road,
                speed=25,
                lane_id=self.config["initial_lane_id"],
                spacing=self.config["ego_spacing"]
            )
            vehicle = self.action_type.vehicle_class(self.road, vehicle.position, vehicle.heading, vehicle.speed)
            self.controlled_vehicles.append(vehicle)
            self.road.vehicles.append(vehicle)

            for _ in range(others):
                vehicle = other_vehicles_type.create_random(self.road, spacing=1 / self.config["vehicles_density"])
                vehicle.randomize_behavior()
                self.road.vehicles.append(vehicle)
        if self.config['save_positions']:
            np.save('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy', self.road.vehicles)
        if self.config['road_static']:
            self.road.vehicles = np.load('save_positions/save_positions_' + str(self.config["lanes_count"]) + '.npy',
                                         allow_pickle=True)
            self.controlled_vehicles[0] = self.road.vehicles[0]

        # self.save_road = static_V(self.road.vehicles)
        # print(self.save_road)

    def mpc_observation(self):
        # self.mpc_env.decision = self.observation_type.decision
        # self.mpc_env.target_lane = self.observation_type.target_lane
        self.mpc_obs = self.mpc_env.observe()
        self.mpc_action = mpc.make_step(self.mpc_obs)
        self.mpc_action = self.mpc_action.reshape(1, 2)[0]
        # u0 = copy.deepcopy(self.mpc_action)
        self.mpc_action[0] = utils.lmap(self.mpc_action[0], self.config['action']['acceleration_range'], [-1, 1])
        self.mpc_action[1] = utils.lmap(self.mpc_action[1], self.config['action']['steering_range'], [-1, 1])

        # return mpc_obs
        return None

    # def mpc_reward(self):
    #     if self.config['mpc_path'] is None:
    #         reward_a = -math.sqrt((self.mpc_action[1] - self.controlled_vehicles[0].action['steering']) ** 2) / (
    #                 np.pi / 10 * 2) + 1
    #         reward_s = -math.sqrt((self.mpc_action[0] - self.controlled_vehicles[0].action['acceleration']) ** 2) / (5 * 2) + 1
    #         reward = reward_s + reward_a
    #     else:
    #         #['pos x', 'pos y', 'steering', 'acc', 'speed', 'solver', 'target y', 'heading','distance','input1','input2']
    #         reward_a = -math.sqrt((self.ref_path[self.steps][3] - self.controlled_vehicles[0].action['acceleration']) ** 2) / (
    #                 np.pi / 10 * 2) + 1
    #         reward_s = -math.sqrt((self.ref_path[self.steps][3] - self.controlled_vehicles[0].action['steering']) ** 2) / (
    #                     5 * 2) + 1
    #
    #
    #     return reward

    def _reward(self, action: Action) -> float:
        """
        The reward is defined to foster driving at high speed, on the rightmost lanes, and to avoid collisions.
        :param action: the last action performed
        :return: the corresponding reward
        """
        if self.config['train']:
            EGO_POS = np.append(self.controlled_vehicles[0].position,
                                [self.controlled_vehicles[0].speed, self.controlled_vehicles[0].lane_index[2]])

            if self.config['mpc_path'] is None:
                # parm_t, parm_s, parm_xx, parm_yy = 0.25, 0.25, 0.25, 0.25
                parm_t, parm_s, parm_xx, parm_yy = 0.4, 0.4, 0.1, 0.1
                # parm_t, parm_s, parm_xx, parm_yy = 0.5, 0.5, 0, 0

                reward_xx = 1 - min(np.abs(self.steer_pre - self.controlled_vehicles[0].action['steering']),0.2)/0.2
                reward_yy = 1 - min(np.abs(self.acc_pre - self.controlled_vehicles[0].action['acceleration']),1.5)/1.5

                reward_t =max(-1 / (0.25 * 4) * np.abs(self.vehicle.position[1] - 4 * self.target_lane) + 1, 0)

                if 25<= self.controlled_vehicles[0].speed <=30:
                    reward_s = 1 - (30 - self.controlled_vehicles[0].speed) / 5
                else:
                    reward_s = 0

                reward = parm_t * reward_t + parm_s * reward_s + parm_xx * reward_xx + parm_yy * reward_yy
                # self.mpc_obs_next = self.mpc_observation()

                self.steer_pre = self.controlled_vehicles[0].action['steering']
                self.acc_pre = self.controlled_vehicles[0].action['acceleration']
            else:
                pass
            # reward = self.mpc_reward()

            reward = self.config["collision_reward"] if self.vehicle.crashed or \
                     (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                     self.vehicle.position[0] < self.trajectory[0][0] or \
                     self.controlled_vehicles[0].speed <= self.config['stop_speed'] else reward

        else:
            reward = 0

        return reward

    def _is_terminal(self) -> bool:
        """The episode is over if the ego vehicle crashed or the time is out."""

        isdone = self.vehicle.crashed or \
                 self.steps >= self.config["duration"] or \
                 (self.config["offroad_terminal"] and not self.vehicle.on_road) or \
                 self.vehicle.position[0] < self.trajectory[0][0] or \
                 self.controlled_vehicles[0].speed <= self.config['stop_speed']
                 # self.controlled_vehicles[0].lane_index[2] not in self.allowed_lane or \


        return isdone

    def _cost(self, action: int) -> float:
        """The cost signal is the occurrence of collision."""
        return float(self.vehicle.crashed)

def static_V(array):
    save_road = []
    for k in array:
        save_road.append(k)
    return save_road

def static_vars(**kwargs):
    def decorate(func):
        for k in kwargs:
            setattr(func, k, kwargs[k])
        return func
    return decorate

register(
    id='highway-v0',
    entry_point='highway_env.envs:HighwayEnv',
)

register(
    id='highway-fast-v0',
    entry_point='highway_env.envs:HighwayEnvFast',
)

register(
    id='myhighway-v0',
    entry_point='highway_env.envs:MyHighwayEnv',
)

register(
    id='myhighway-v1',
    entry_point='highway_env.envs:MyHighwayEnv_LaneDone',
)

register(
    id='myhighway-v2',
    entry_point='highway_env.envs:MyHighwayEnv_Kinematic',
)

register(
    id='myhighway-v3',
    entry_point='highway_env.envs:MyHighwayEnv_Kinematic2',
)

register(
    id='myhighway-v4',
    entry_point='highway_env.envs:MyHighwayEnv_AugularV1',
)

register(
    id='myhighway-v5',
    entry_point='highway_env.envs:HighwayEnv_Main',
)

register(
    id='myhighway-v6',
    entry_point='highway_env.envs:HighwayEnv_Simple',
)

register(
    id='myhighway-v7',
    entry_point='highway_env.envs:HighwayEnv_Long',
)


register(
    id='refhighway-v0',
    entry_point='highway_env.envs:HighwayEnv_Reference_Reward',
)

register(
    id='refhighway-v1',
    entry_point='highway_env.envs:HighwayEnv_MPC_Buffer',
)